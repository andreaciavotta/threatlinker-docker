# Create your views here.
from django.shortcuts import render, redirect
from django.contrib import messages
from core.functions.cve_handling import validate_cve_list
from data.functions.handle_cve import get_or_create_cves
from data.models import CVE, CAPEC
from core.models import Task, SingleCorrelation
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from core.tasks.main import process_task_with_chord
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
import pandas as pd
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def homepage(request):
    return render(request, 'homepage.html')

def view_cve(request, cve_id):
    try:
        cve = CVE.objects.get(id=cve_id)
    except CVE.DoesNotExist:
        # Chiamata alla view_error_page con il messaggio personalizzato
        return view_error_page(request, f'{cve_id} not found.')
 
    # Prepare the context to pass the CVE data to the template
    context = {
        'cve': cve
    }

    # Render the 'cve.html' template and pass the CVE context
    return render(request, 'view/cve/view_cve.html', context)

def view_capec(request, capec_id):
    # Fetch the CVE from the database based on the ID
    try:
        capec = CAPEC.objects.get(id=capec_id)
    except CAPEC.DoesNotExist:
        # Chiamata alla view_error_page con il messaggio personalizzato
        return view_error_page(request, f'{capec_id} not found.')
 
    # Prepara il contesto da passare al template
    context = {
        'capec': capec
    }

    # Renderizza il template 'view/capec.html' e passa il contesto del CAPEC
    return render(request, 'view/capec/view_capec.html', context)

def view_error_page(request, message=None):
    """
    Rende la pagina di errore con un messaggio personalizzato (se fornito).
    Se non viene fornito nessun messaggio, viene utilizzato un messaggio di default.
    :param request: La richiesta HTTP
    :param message: Il messaggio di errore da visualizzare (opzionale)
    :return: La risposta HTTP con il template renderizzato
    """
    if not message:
        message = "Sorry, an error occurred. Please try again later."
    
    return render(request, 'errors/error_page.html', {'message': message})

### Task
def create_task(request):
    similarity_methods = ["SBERT Hyb", "ATTACKBERT Hyb"]
    
    if request.method == "POST":
        cve_list = request.POST.get("cve_list", "").strip()
        uploaded_file = request.FILES.get("file")
        selected_method = request.POST.get("similarity_methods")
        print(f"Selected method: {selected_method}")
        
        if not cve_list and not uploaded_file:
            messages.error(request, "You must provide either a list of CVEs or upload a file.")
            return redirect("vista:create_task")

        if not selected_method:
            messages.error(request, "At least one similarity method must be selected.")
            return redirect("vista:create_task")

        # Validate CVE IDs
        cve_ids = validate_cve_list(cve_list)
        
        if cve_list and not cve_ids:
            messages.error(request, "Invalid CVE format. Use CVE-YYYY-XXXXX (e.g., CVE-2023-12345).")
            return redirect("vista:create_task")

        # Fetch CVE details from NVD database
        cve_data = get_or_create_cves(cve_ids)
        
        # Mostra il riepilogo nella pagina
        return render(request, "tasks/make_request.html", {
            "similarity_methods": similarity_methods,
            "cve_count": len(cve_data),
            "cve_data": cve_data,
            "selected_method": selected_method
        })

    return render(request, "tasks/create_task.html", {"similarity_methods": similarity_methods})


@csrf_exempt  # Solo se usi fetch e non hai CSRF nei header
def start_task(request):
    if request.method == "POST":
        try:
            # Ottieni i dati dal form normale (non JSON)
            task_name = request.POST.get("task_name")
            task_description = request.POST.get("task_description")
            # Qui, invece di usare get() che potrebbe non funzionare come previsto, usiamo getlist()
            similarity_method_list = request.POST.getlist("similarity_method")
            # Se la lista non è vuota, prendi il primo elemento e applica strip()
            similarity_method = similarity_method_list[0].strip() if similarity_method_list else ""
            
            print(f"Metodo: {similarity_method}")
            cve_list = request.POST.getlist("cve_list")  # Ottiene una lista se è multiplo
            
            if not task_name or not task_description or not similarity_method or not cve_list:
                return JsonResponse({"success": False, "message": "All fields are required."}, status=400)

            # Creazione della task
            task = Task.objects.create(
                name=task_name,
                notes=task_description,
                ai_models=similarity_method,
                cve_hosts=",".join(cve_list)  # Concatena in una stringa separata da virgole
            )

            task.save()  # Salva il dizionario cve_hosts nella Task

            task_id = task.id  # 🔥 Prendi l'ID della task appena creata

            task.save()  # Salva la task nel database

            # Avvia il Chord Celery con l'ID della task
            chord_task = process_task_with_chord.delay(task_id)
            # Messaggio di successo
            messages.success(request, 'Correlation Task started successfully!')

            # Reindirizzamento alla pagina di dettaglio della Task
            return redirect('vista:tasks_list')

        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)}, status=500)

    return JsonResponse({"success": False, "message": "Invalid request."}, status=400)


### Task View

def tasks_list(request):
    tasks = Task.objects.all()  # Recupera tutte le Task
    return render(request, 'tasks/tasks_list.html', {'tasks': tasks})

# Vista per eliminare una Task
def delete_task(request, task_id):
    if request.method == 'POST':
        try:
            task = Task.objects.get(id=task_id)
            # Ora puoi eliminare la Task
            task.delete()
            return redirect('vista:tasks_list')  # Redirect alla lista delle task
        except Task.DoesNotExist:
            return HttpResponse("Task not found", status=404)
    else:
        return HttpResponse("Invalid request method", status=400)
    
def task_detail(request, task_id):
    """
    Mostra i dettagli di una Task e le CVE correlate.
    """
    # Recupera la Task
    task = get_object_or_404(Task, id=task_id)
    
    # Recupera le SingleCorrelation associate alla Task
    correlations = SingleCorrelation.objects.filter(task=task)

    # Recupera le CVE dal database
    cve_details = []
    for correlation in correlations:
        cve = CVE.objects.filter(id=correlation.cve_id).first()  # Trova la CVE corrispondente
        cve_details.append({
            "cve_id": correlation.cve_id,
            "description": cve.description if cve else "No description available",
        })

    context = {
        "task": task,
        "cve_details": cve_details,
    }
    
    return render(request, "tasks/task_detail.html", context)


def single_correlation_detail(request, task_id, cve_id):
    """
    Mostra i dettagli della correlazione tra una CVE e le CAPEC, con la classifica per il modello "_keyword".
    """
    # Recupera la Task
    task = get_object_or_404(Task, id=task_id)

    # Recupera la correlazione per la CVE specificata
    correlation = get_object_or_404(SingleCorrelation, task=task, cve_id=cve_id)

    # Recupera la CVE dal database
    cve = get_object_or_404(CVE, id=cve_id)

    # Filtra solo i risultati che contengono "_keyword"
    keyword_models = {key: value for key, value in correlation.similarity_scores.items() if "_keyword" in key}

    if not keyword_models:
        return render(request, "tasks/single_correlation_detail.html", {
            "task": task,
            "cve": cve,
            "correlation": correlation,
            "no_data": True
        })

    # Prendi solo il primo modello (dovrebbe essercene solo uno con _keyword)
    model_name, capec_scores = list(keyword_models.items())[0]
    
    # Converte il nome del modello da "{MODELLO}_keyword" a "{MODELLO} Hyb"
    formatted_model_name = model_name.replace("_keyword", " Hyb")

    # Ordina le CAPEC per final_score
    capec_ranking = sorted(capec_scores, key=lambda x: x[1].get("final_score", 0), reverse=True)

    # Recupera i dettagli delle CAPEC associate
    capec_details = {capec.id: capec.name for capec in CAPEC.objects.filter(id__in=[capec_id for capec_id, _ in capec_ranking])}

    context = {
        "task": task,
        "cve": cve,
        "correlation": correlation,
        "formatted_model_name": formatted_model_name,
        "capec_ranking": capec_ranking,
        "capec_details": capec_details,
    }

    return render(request, "tasks/single_correlation_detail.html", context)

# Excel Results

def export_task_excel(request, task_id):
    """
    Esporta un file Excel contenente i dettagli della task e le correlazioni CVE-CAPEC,
    con colori e link cliccabili direttamente.
    """
    task = get_object_or_404(Task, id=task_id)
    correlations = SingleCorrelation.objects.filter(task=task)
    
    # Creazione file Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Task_{task.id}"

    # Definizione colori
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Grigio chiaro
    cve_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Giallo chiaro
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde chiaro
    orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # Arancione

    # Creazione intestazione
    headers = ["CVE ID", "CVE Description"] + [f"Rank {i+1}" for i in range(10)]
    ws.append(headers)

    # Applica colore alle intestazioni
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).fill = header_fill

    # Riempimento dati
    for correlation in correlations:
        cve = get_object_or_404(CVE, id=correlation.cve_id)
        capec_rankings = []
        
        # Prendiamo solo il primo modello con "_keyword"
        keyword_models = {key: value for key, value in correlation.similarity_scores.items() if "_keyword" in key}
        if keyword_models:
            model_name, capec_scores = list(keyword_models.items())[0]

            # Ordina le CAPEC per final_score in ordine decrescente
            sorted_capecs = sorted(capec_scores, key=lambda x: x[1].get("final_score", 0), reverse=True)[:10]

            # Creiamo le prime 10 CAPEC con link
            for idx in range(10):
                if idx < len(sorted_capecs):
                    capec_id_with_prefix, capec_data = sorted_capecs[idx]
                    
                    # Estrai solo il numero dell'ID rimuovendo "CAPEC-"
                    capec_number = capec_id_with_prefix.replace("CAPEC-", "")
                    
                    capec_name = get_object_or_404(CAPEC, id=capec_id_with_prefix).name
                    capec_link = f"https://capec.mitre.org/data/definitions/{capec_number}.html"
                    capec_rankings.append((f"{capec_id_with_prefix} : {capec_name}", capec_link))
                else:
                    capec_rankings.append(("-", ""))  # Se non ci sono abbastanza CAPEC, riempi con "-"

        # Aggiunta riga con CVE e CAPEC
        row = [
            (cve.id, f"https://nvd.nist.gov/vuln/detail/{cve.id}"),  # CVE ID con link
            (cve.description, ""),  # Descrizione senza link
            *capec_rankings  # CAPEC con link
        ]
        
        # Scrittura nel foglio Excel
        row_num = ws.max_row + 1
        for col_num, (text, link) in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=text)
            
            # Aggiunta di link se presenti
            if link:
                cell.hyperlink = link
                cell.style = "Hyperlink"

            # Colorazione celle
            if col_num == 1 or col_num == 2:  # CVE ID e Descrizione
                cell.fill = cve_fill
            elif col_num >= 3:
                if col_num - 2 <= 5:  # Rank 1-5 → Verde
                    cell.fill = green_fill
                else:  # Rank 6-10 → Arancione
                    cell.fill = orange_fill

    # Salvataggio file temporaneo
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"task_{task.id}_{timestamp}.xlsx"
    file_path = os.path.join("/tmp", filename)
    wb.save(file_path)

    # Invio come risposta HTTP
    with open(file_path, "rb") as excel_file:
        response = HttpResponse(excel_file.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f"attachment; filename={filename}"

    return response
